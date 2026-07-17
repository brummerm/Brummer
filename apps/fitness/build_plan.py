#!/usr/bin/env python3
"""Parse Workout_Plan.xlsx into app/plan.json.

Re-run whenever the spreadsheet changes:
    python3 build_plan.py [path/to/Workout_Plan.xlsx]
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else str(Path.home() / "Downloads" / "Workout_Plan.xlsx")
OUT = Path(__file__).parent / "static" / "plan.json"

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

PHASE_SHEETS = [
    ("Phase 1 - Foundation (Mo 1-3)", 1),
    ("Phase 2 - Strength (Mo 4-6)", 2),
    ("Phase 3 - Engine (Mo 7-9)", 3),
    ("Phase 4 - Peak (Mo 10-12)", 4),
]

WEEK_RULE_RE = re.compile(r"^Wk\s*(\d+)\s*(?:-\s*(\d+))?\s*:\s*(.+)$", re.S)
PARITY_RULE_RE = re.compile(r"^(Odd|Even|Most)\s+weeks:\s*(.+)$", re.S)
NAME_RX_RE = re.compile(r"^(.*?)\s+—\s+(.+)$", re.S)  # "Name — prescription"


def clean(v):
    if v is None:
        return None
    s = str(v).replace(" ", " ").strip()
    return s if s else None


def parse_prescription(rx):
    """Best-effort structure from strings like '4 x 8', '3 x 45 sec', '3 x 40 yd/side',
    '4 x (max minus 1)', '2 x 12 each', '3 x 10/leg', '5 x 3-5', '3 x 8 min (10-15% grade)'."""
    out = {"raw": rx}
    m = re.match(r"^(\d+)\s*x\s*(.+)$", rx, re.S)
    if not m:
        return out
    out["sets"] = int(m.group(1))
    rest = m.group(2).strip()

    per = re.search(r"/\s*(side|leg|arm)\b|\beach\b", rest)
    if per:
        out["perSide"] = True

    t = re.match(r"^(\d+)(?:\s*-\s*(\d+))?\s*(sec|min)\b", rest)
    if t:
        out["type"] = "time"
        out["seconds"] = int(t.group(1)) * (60 if t.group(3) == "min" else 1)
        if t.group(2):
            out["secondsHigh"] = int(t.group(2)) * (60 if t.group(3) == "min" else 1)
        return out

    d = re.match(r"^(\d+)\s*(yd|mi|km)\b", rest)
    if d:
        out["type"] = "distance"
        out["distance"] = int(d.group(1))
        out["unit"] = d.group(2)
        return out

    iv = re.match(r"^(\d+)\s*m\b", rest)
    if iv:
        out["type"] = "interval"
        out["distance"] = int(iv.group(1))
        out["unit"] = "m"
        return out

    if "max" in rest.lower():
        out["type"] = "max"
        out["target"] = rest
        return out

    r = re.match(r"^(\d+)(?:\s*-\s*(\d+))?", rest)
    if r:
        out["type"] = "reps"
        out["reps"] = int(r.group(1))
        if r.group(2):
            out["repsHigh"] = int(r.group(2))
        return out

    out["type"] = "other"
    out["target"] = rest
    return out


def parse_item(text):
    """One cell of a phase-day column -> item dict."""
    wr = WEEK_RULE_RE.match(text)
    if wr:
        lo = int(wr.group(1))
        hi = int(wr.group(2)) if wr.group(2) else lo
        return {"kind": "weekRule", "weeks": [lo, hi], "text": wr.group(3).strip(), "raw": text}

    pr = PARITY_RULE_RE.match(text)
    if pr:
        which = pr.group(1).lower()  # odd | even | most
        return {"kind": "weekRule", "parity": "default" if which == "most" else which,
                "text": pr.group(2).strip(), "raw": text}

    nm = NAME_RX_RE.match(text)
    if nm:
        item = {"kind": "exercise", "name": nm.group(1).strip()}
        item.update(parse_prescription(nm.group(2).strip()))
        return item

    return {"kind": "note", "raw": text}


def parse_phase(ws, phase_num):
    title = clean(ws.cell(1, 1).value) or ""
    desc = clean(ws.cell(2, 1).value) or ""
    wk = re.search(r"WEEKS\s+(\d+)\s*-\s*(\d+)", title, re.I)
    weeks = [int(wk.group(1)), int(wk.group(2))] if wk else [None, None]

    days = []
    for col in range(1, 8):
        header = clean(ws.cell(5, col).value) or ""
        parts = header.split("\n")
        session_title = parts[0].strip()
        duration = None
        if len(parts) > 1:
            dm = re.search(r"\(([^)]+)\)", parts[1])
            duration = dm.group(1) if dm else parts[1].strip()

        raw_items, notes = [], None
        for row in range(6, 25):
            v = clean(ws.cell(row, col).value)
            if not v:
                continue
            if v.startswith("Notes:"):
                notes = v[len("Notes:"):].strip()
                continue
            raw_items.append(v)

        # Merge consecutive week-rule lines into one week-variant item.
        items = []
        for raw in raw_items:
            item = parse_item(raw)
            if item["kind"] == "weekRule":
                variant = {k: item[k] for k in ("weeks", "parity", "text") if k in item}
                if items and items[-1]["kind"] == "weekly":
                    items[-1]["variants"].append(variant)
                else:
                    items.append({"kind": "weekly", "variants": [variant]})
            else:
                items.append(item)

        days.append({
            "day": DAYS[col - 1],
            "title": session_title,
            "duration": duration,
            "items": items,
            "notes": notes,
        })

    return {"phase": phase_num, "name": title, "description": desc,
            "weeks": weeks, "days": days}


def parse_run_ruck(ws):
    weeks = []
    for row in range(5, 60):
        w = ws.cell(row, 1).value
        if w is None:
            continue
        notes = clean(ws.cell(row, 7).value) or ""
        weeks.append({
            "week": int(w),
            "phaseName": clean(ws.cell(row, 2).value),
            "runMiles": ws.cell(row, 3).value,
            "keyRun": clean(ws.cell(row, 4).value),
            "ruckMiles": clean(ws.cell(row, 5).value),
            "ruckWeight": clean(ws.cell(row, 6).value),
            "notes": notes,
            "deload": "DELOAD" in notes.upper(),
            "benchmark": "BENCHMARK" in notes.upper(),
            "taper": "TAPER" in notes.upper(),
        })
    return weeks


def parse_mobility(ws):
    routines, current = [], None
    for row in range(5, 40):
        routine = clean(ws.cell(row, 1).value)
        exercise = clean(ws.cell(row, 2).value)
        if not exercise:
            continue
        if routine:
            current = {"name": routine, "exercises": []}
            routines.append(current)
        if current is None:
            continue
        current["exercises"].append({
            "name": exercise,
            "sets": clean(ws.cell(row, 3).value),
            "duration": clean(ws.cell(row, 4).value),
            "notes": clean(ws.cell(row, 5).value),
        })
    return routines


def parse_kv_sheet(ws, max_row=60):
    """Overview-style sheets: label col A, text col B, section headers when col B empty."""
    sections, current = [], None
    for row in range(1, max_row):
        a, b = clean(ws.cell(row, 1).value), clean(ws.cell(row, 2).value)
        if a and not b:
            current = {"heading": a, "entries": []}
            sections.append(current)
        elif a and b:
            if current is None:
                current = {"heading": "", "entries": []}
                sections.append(current)
            current["entries"].append({"label": a, "text": b})
    return sections


def parse_meals(ws):
    meals, current = [], None
    totals = []
    for row in range(5, 40):
        meal = clean(ws.cell(row, 1).value)
        option = clean(ws.cell(row, 2).value)
        foods = clean(ws.cell(row, 3).value)
        if not foods:
            continue
        if foods.startswith(("PLAN TOTAL", "DAILY TARGET")):
            totals.append({
                "label": foods,
                "protein": clean(ws.cell(row, 4).value),
                "carbs": clean(ws.cell(row, 5).value),
                "fat": clean(ws.cell(row, 6).value),
                "calories": clean(ws.cell(row, 7).value),
            })
            continue
        if meal:
            current = {"slot": meal, "options": []}
            meals.append(current)
        if current is None or not option:
            continue

        def num(c):
            v = ws.cell(row, c).value
            return float(v) if isinstance(v, (int, float)) else clean(v)

        current["options"].append({
            "option": option,
            "foods": foods,
            "protein": num(4), "carbs": num(5), "fat": num(6), "calories": num(7),
            "notes": clean(ws.cell(row, 8).value),
        })
    return meals, totals


def parse_grocery(ws):
    items, category = [], None
    for row in range(5, 40):
        cat = clean(ws.cell(row, 1).value)
        item = clean(ws.cell(row, 2).value)
        if not item:
            continue
        if cat:
            category = cat
        items.append({
            "category": category,
            "item": item,
            "amount": clean(ws.cell(row, 3).value),
            "notes": clean(ws.cell(row, 4).value),
        })
    return items


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    plan = {
        "meta": {
            "title": clean(wb["Overview & How to Use"].cell(1, 1).value),
            "subtitle": clean(wb["Overview & How to Use"].cell(2, 1).value),
            "totalWeeks": 52,
            "source": Path(XLSX).name,
        },
        "overview": parse_kv_sheet(wb["Overview & How to Use"]),
        "phases": [parse_phase(wb[s], n) for s, n in PHASE_SHEETS],
        "runRuck": parse_run_ruck(wb["Run & Ruck Progression"]),
        "mobility": parse_mobility(wb["Mobility & Flexibility"]),
        "nutrition": {
            "overview": parse_kv_sheet(wb["Nutrition Overview"]),
            "grocery": parse_grocery(wb["Grocery Staples"]),
        },
    }
    meals, totals = parse_meals(wb["Daily Meal Plan"])
    plan["nutrition"]["meals"] = meals
    plan["nutrition"]["mealTotals"] = totals

    OUT.parent.mkdir(parents=True, exist_ok=True)
    # Scrub program-name branding from display text.
    out_json = json.dumps(plan, indent=1)
    out_json = out_json.replace("SFAS PREPARATION PROGRAM", "TRAINING PROGRAM")
    out_json = out_json.replace("SFAS-specific", "event-specific")
    out_json = out_json.replace("SFAS", "selection")
    OUT.write_text(out_json)
    print(f"Wrote {OUT} ({OUT.stat().st_size // 1024} KB)")
    print(f"Phases: {[p['weeks'] for p in plan['phases']]}")
    print(f"Run/ruck weeks: {len(plan['runRuck'])}")
    print(f"Mobility routines: {len(plan['mobility'])}")
    print(f"Meal slots: {len(meals)}")
    print(f"Grocery items: {len(plan['nutrition']['grocery'])}")


if __name__ == "__main__":
    main()
